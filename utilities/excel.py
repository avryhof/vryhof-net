import csv
import os
import zipfile

from django.forms import model_to_dict
from openpyxl import load_workbook, Workbook
from pyxlsb import open_workbook

from utilities.api_utils import to_dict, strip_dict
from utilities.debugging import log_message


def list_worksheets(excel_file):
    sheets = []
    file_name, file_extension = os.path.splitext(excel_file)

    if os.path.exists(excel_file):
        if file_extension.lower() == ".xlsb":
            try:
                wb = open_workbook(excel_file)
            except Exception as e:
                print("BAD File: {}".format(excel_file))
                print(e)
            else:
                sheets = wb.sheets
        else:
            try:
                wb = load_workbook(filename=excel_file, data_only=True, read_only=True)
            except zipfile.BadZipFile:
                print("Bad Zip File")
                print(excel_file)
            else:
                sheets = wb.sheetnames

    return sheets


def get_worksheet(excel_file, worksheet_name=False):
    """
    Returns an OpenPyxl (or pyxlsb) worksheet object as named in worksheet name.
    If worksheet_name is not specified, the first worksheet in the workbook is returned.

    :param excel_file: the filename for a Microsoft Excel (xls, or xlsx) file.
    :param worksheet_name: The name of a worksheet (tab) in the workbook. If omitted, the first worksheet is returned.
    :return:
        An OpenPyxl worksheet object.
    """
    worksheet = []

    file_name, file_extension = os.path.splitext(excel_file)

    if os.path.exists(excel_file):
        if file_extension.lower() == ".xlsb":
            try:
                wb = open_workbook(excel_file)
            except Exception as e:
                print("BAD File: {}".format(excel_file))
                print(e)
            else:
                if worksheet_name:
                    worksheet = wb.get_sheet(worksheet_name)
                else:
                    worksheet = wb.get_sheet(1)

        else:
            try:
                wb = load_workbook(filename=excel_file, data_only=True, read_only=True)
            except zipfile.BadZipFile:
                print("Bad Zip File")
                print(excel_file)
            else:
                worksheet = wb.worksheets[0]
                if worksheet_name:
                    for ws in wb.worksheets:
                        if ws.title == worksheet_name:
                            worksheet = ws
                            break

    return worksheet


def csv_to_dicts(csv_file, **kwargs):
    headers = kwargs.get("headers", False)
    encoding = kwargs.pop("encoding", "utf8")
    primary_key = kwargs.get("primary_key", False)
    fix_keys = kwargs.pop("fix_keys", False)

    sheet = []

    df = open(csv_file, "r", encoding=encoding)

    if not headers:
        rows = csv.DictReader(df, **kwargs)
    else:
        fieldnames = kwargs.pop("headers")
        kwargs.update(dict(fieldnames=fieldnames))
        rows = csv.DictReader(df, **kwargs)

    for row in rows:
        if not primary_key or (primary_key and row.get(primary_key) and row.get(primary_key) != ""):
            if not isinstance(row, dict):
                row = to_dict(row)

            if fix_keys:
                row = strip_dict(row)

            sheet.append(row)

    df.close()

    return sheet


def write_dicts_to_csv(dicts, csv_file):
    """
    Writes a list of dicts out to a csv file.
    Field names are auto-detected from the keys of the first dict in the list.
    :param dicts: A list of dicts to be written
    :param csv_file: the filename for a target csv file.
    :return:
    """
    keys = list(dicts[0].keys())

    try:
        with open(csv_file, "w", encoding="utf8", newline="") as cf:
            cw = csv.DictWriter(cf, fieldnames=keys)
            cw.writeheader()
            cw.writerows(dicts)
            cf.close()

    except Exception as e:
        was_written = False

    else:
        was_written = True

    return was_written


def convert_csv_to_xlsx(csv_file, xlsx_file):
    wb = Workbook()
    ws = wb.active
    with open(csv_file, "r") as f:
        for row in csv.reader(f):
            ws.append(row)

    try:
        wb.save(xlsx_file)
    except Exception as e:
        log_message(e)
        retn = False
    else:
        retn = True

    return retn


def write_dicts_to_xlsx(dicts, xlsx_file):
    retn = True

    csv_file = xlsx_file.replace(".xlsx", ".csv")

    if write_dicts_to_csv(dicts, csv_file):
        if convert_csv_to_xlsx(csv_file, xlsx_file):
            os.remove(csv_file)
        else:
            retn = False
    else:
        retn = False

    return retn


def excel_to_dicts(excel_file, **kwargs):
    """
    Returns a worksheet within an Excel workbook as a list of dicts.  If target_worksheet is omitted, the first
    worksheet is returned.
    :param excel_file: The filename for a Microsoft Excel file
    :param kwargs:
        key_row - Sometimes an Excel sheet will have other data and headers before the tabular data.  This specifies
                  which worksheet row to start reading tabular data from.
        target_worksheet - The name of the worksheet (tab) to get data from.
    :return:
        A list of dicts containing the tabular data from the specified worksheet.
    """
    headers = kwargs.get("headers", False)
    key_row = kwargs.get("key_row", 1)
    target_worksheet = kwargs.get("worksheet", False)

    sheet = []

    file_name, file_extension = os.path.splitext(excel_file)

    if os.path.exists(excel_file):
        worksheet = get_worksheet(excel_file, target_worksheet)

        sheet = []
        keys = []

        if file_extension.lower() == ".xlsb":
            i = 0

            if headers:
                keys = headers

            for row in worksheet.rows():
                if not headers and i == key_row - 1:
                    for col in row:
                        keys.append(col.v)
                    headers = keys

                i = i + 1

                values = []
                for column in row:
                    values.append(column.v)
                if values[0]:
                    sheet.append(dict(list(zip(keys, values))))

        else:
            if not headers:
                dict_keys_row = worksheet.iter_rows(min_row=key_row, max_row=key_row)
                for dict_keys in dict_keys_row:
                    for dict_key in dict_keys:
                        keys.append(dict_key.value)

            elif isinstance(headers, list):
                keys = headers

            new_keys = []
            for key in keys:
                if isinstance(key, str):
                    new_keys.append(key.strip())
                else:
                    new_keys.append(key)
            keys = new_keys

            rows = worksheet.iter_rows(min_row=(key_row + 1))
            for row in rows:
                values = []
                for column in row:
                    values.append(column.value)
                if values[0]:
                    sheet.append(dict(list(zip(keys, values))))

    return sheet


def excel_to_csv(excel_file, **kwargs):
    """
    Converts an Excel file to a CSV file with the same base filename, in the same location.
    :param excel_file:
    :param kwargs:
    :return:
    """
    key_row = kwargs.get("key_row", 1)
    target_worksheet = kwargs.get("worksheet", False)
    target_filename = kwargs.get("output", False)

    csv_filename = False

    if os.path.exists(excel_file):
        if target_filename:
            csv_file = target_filename
        else:
            path_parts = os.path.split(excel_file)
            excel_file_name = path_parts[-1]
            excel_file_extension = ".%s" % excel_file_name.split(".")[-1]
            csv_file_name = excel_file_name.replace(excel_file_extension, ".csv")
            csv_file = excel_file.replace(excel_file_name, csv_file_name)

        sheet_dict = excel_to_dicts(excel_file, key_row=key_row, target_worksheet=target_worksheet)
        write_dicts_to_csv(sheet_dict, csv_file)
        csv_filename = csv_file

    return csv_filename


def dicts_to_table(dicts):
    """
    Converts a list of dicts into something that can easily be displayed as a table.
    :param dicts: A list of dicts to be written
    :return:
    """
    retn = dict()

    keys = list(dicts[0].keys())
    data = []

    try:
        for dict_item in dicts:
            data.append(list(dict_item.values()))

    except Exception:
        retn = False

    else:
        retn = dict(keys=keys, data=data)

    return retn


def model_to_csv(model, columns, csv_file):
    try:
        with open(csv_file, "w", encoding="utf8", newline="") as cf:
            cw = csv.DictWriter(cf, fieldnames=columns)
            cw.writeheader()
            for obj in model.objects.all():
                csv_dict = model_to_dict(obj, fields=columns)
                cw.writerow(csv_dict)
            cf.close()
    except Exception as e:
        was_written = False
    else:
        was_written = True

    return was_written
